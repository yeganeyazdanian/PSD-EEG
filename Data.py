import numpy as np
import matplotlib.pyplot as plt
import openpyxl as exc
from scipy.signal import iirfilter, welch, lfilter
import math
import mne
import re


# import only system from os
from os import system, name
# import sleep to show output for some time period
from time import sleep
# define our clear function


def clear():
    sleep(0.001)
    # for windows
    if name == 'nt':
        _ = system('cls')
    # for mac and linux(here, os.name is 'posix')
    else:
        _ = system('clear')


# Required input defintions are as follows;
# time:   Time between samples
# band:   The bandwidth around the centerline freqency that you wish to filter
# freq:   The centerline frequency to be filtered
# ripple: The maximum passband ripple that is allowed in db
# order:  The filter order.  For FIR notch filters this is best set to 2 or 3,
#         IIR filters are best suited for high values of order.  This algorithm
#         is hard coded to FIR filters
# filter_type: 'butter', 'bessel', 'cheby1', 'cheby2', 'ellip'
# data:         the data to be filtered


def Implement_Notch_Filter(fs, band, freq, ripple, order, filter_type, data):
    nyq = fs/2.0
    low = freq - band/2.0
    high = freq + band/2.0
    low = low/nyq
    high = high/nyq
    b, a = iirfilter(order, [low, high], rp=ripple, btype='bandstop',
                     analog=False, ftype=filter_type)
    filtered_data = lfilter(b, a, data)
    return filtered_data


# get data from edf file


def Get_EDF_data():
    f = mne.io.read_raw_edf("Bae_1.edf")

    n = f.get_data()
    n2 = np.reshape(n, (169088, 40))
    data = Implement_Notch_Filter(256, 2, 50, 5, 2, 'butter', n2)
    return data

# get data from xlsx file to numpy array


def Get_xlsx_data(row_start, row_end, column1, column2, column3):

    # imported xlsx data
    wb = exc.load_workbook('fnirs-tums.xlsx')
    sheet = wb['Export 1']
    x = np.zeros((row_end-row_start, 3))
    for i in range(0, row_end-row_start):
        x[i, 0] = sheet.cell(i+row_start, column1).value
        x[i, 1] = sheet.cell(i+row_start, column2).value
        x[i, 2] = sheet.cell(i+row_start, column3).value
    return x

# get datat from txt file


def Get_txt_data(fp):
    _fp = open(fp+".txt", "r")
    nps = re.split("\t|\n", _fp.read())
    nps.pop()
    nps = nps[0:149000]
    _fp.close()

    m = np.zeros((len(nps)))
    for i in range(0, len(nps)):
        m[i] = float(nps[i])
    return m


def GetABT(freqs, psd_db):
    TStart = np.argwhere(freqs == 4)
    TEnd = np.argwhere(freqs == 8)
    freqT = freqs[TStart[0, 0]:TEnd[0, 0]]
    psd_db_T = psd_db[TStart[0, 0]:TEnd[0, 0]]

    AStart = np.argwhere(freqs == 8)
    AEnd = np.argwhere(freqs == 16)
    freqA = freqs[AStart[0, 0]:AEnd[0, 0]]
    psd_db_A = psd_db[AStart[0, 0]:AEnd[0, 0]]

    BStart = np.argwhere(freqs == 16)
    BEnd = np.argwhere(freqs == 31)
    freqB = freqs[BStart[0, 0]:BEnd[0, 0]]
    psd_db_B = psd_db[BStart[0, 0]:BEnd[0, 0]]

    Theta = np.array([psd_db_T, freqT])
    Beta = np.array([psd_db_B, freqB])
    Alpha = np.array([psd_db_A, freqA])

    data = (Theta, Beta, Alpha)

    return data


def FP(fb):
    freqs, psd = welch(fb[0:], 256, nperseg=1024)
    # freqs, psd = welch(fb[i, 0:], 256, nperseg=1024)
    return (freqs, psd)


def PSDtoDB(psd):
    psd_db = np.zeros((len(psd)))
    for i in range(0, len(psd)):
        psd_db[i] = 10 * math.log(psd[i])
    return psd_db


def Means(param):
    Meanfp1 = np.mean(param[0:][0])
    return Meanfp1


def Mean(param):
    mean = np.mean(param[0:])
    return mean


# the range is 58 to 7075 / and the column range is 2 to 17
# fb = Get_data(58,7075,2,3,4)
# fb = Get_EDF_data()
patient_list = {"Timed": {"fp1": ["patient_1_fp1_T", "patient_2_fp1_T", "patient_3_fp1_T", "patient_4_fp1_T"],
                          "fp2": ["patient_1_fp2_T", "patient_2_fp2_T", "patient_3_fp2_T", "patient_4_fp2_T"],
                          "fpz": ["patient_1_fpz_T", "patient_2_fpz_T", "patient_3_fpz_T", "patient_4_fpz_T"],
                          },
                "UnTimed": {"fp1": ["patient_1_fp1_U", "patient_2_fp1_U", "patient_3_fp1_U", "patient_4_fp1_U"],
                            "fp2": ["patient_1_fp2_U", "patient_2_fp2_U", "patient_3_fp2_U", "patient_4_fp2_U"],
                            "fpz": ["patient_1_fpz_U", "patient_2_fpz_U", "patient_3_fpz_U", "patient_4_fpz_U"],
                            },
                "Relaxed": {"fp1": ["patient_1_fp1_R", "patient_2_fp1_R", "patient_3_fp1_R", "patient_4_fp1_R"],
                            "fp2": ["patient_1_fp2_R", "patient_2_fp2_R", "patient_3_fp2_R", "patient_4_fp2_R"],
                            "fpz": ["patient_1_fpz_R", "patient_2_fpz_R", "patient_3_fpz_R", "patient_4_fpz_R"],
                            }
                }


def Fin_Means():
    means = {}
    # this loop goes through every stage (Timed,Untimed,Relaxed)
    for i in patient_list.items():
        stage, stage_list = i
        alpha_means = []
        beta_means = []
        theta_means = []
        patient_chanels = {"p1": {"alpha": [], "beta": [], "theta": []},
                           "p2": {"alpha": [], "beta": [], "theta": []},
                           "p3": {"alpha": [], "beta": [], "theta": []},
                           "p4": {"alpha": [], "beta": [], "theta": []}}

        percents =        {"fp1": {"alpha": [], "beta": [], "theta": []},
                           "fp2": {"alpha": [], "beta": [], "theta": []},
                           "fpz": {"alpha": [], "beta": [], "theta": []}}
        # this loop goes through every fp in stage (ex index0 = fp1,fp2,fpz for stage Timed)
        for j in stage_list.items():
            fp, fp_stage = j
            # this loop goes through every patient in every fp in stage
            for k in fp_stage:
                clear()
                print("in process: [{}]".format(k))
                fb = Get_txt_data(k)
                freqs, psd = FP(fb)
                psd_db_fp = PSDtoDB(psd)
                theta_fp, beta_fp, alpha_fp = GetABT(freqs, psd_db_fp)
                alphaM = Means(alpha_fp)
                betaM = Means(beta_fp)
                thetaM = Means(theta_fp)
                alpha_means.append(alphaM)
                beta_means.append(betaM)
                theta_means.append(thetaM)
                if "patient_1" in k:
                    patient_chanels["p1"]["alpha"].append(alphaM)
                    patient_chanels["p1"]["beta"].append(betaM)
                    patient_chanels["p1"]["theta"].append(thetaM)
                if "patient_2" in k:
                    patient_chanels["p2"]["alpha"].append(alphaM)
                    patient_chanels["p2"]["beta"].append(betaM)
                    patient_chanels["p2"]["theta"].append(thetaM)
                if "patient_3" in k:
                    patient_chanels["p3"]["alpha"].append(alphaM)
                    patient_chanels["p3"]["beta"].append(betaM)
                    patient_chanels["p3"]["theta"].append(thetaM)
                if "patient_4" in k:
                    patient_chanels["p4"]["alpha"].append(alphaM)
                    patient_chanels["p4"]["beta"].append(betaM)
                    patient_chanels["p4"]["theta"].append(thetaM)
                
                if fp == "fp1":
                    percents["fp1"]["alpha"].append(alphaM)
                    percents["fp1"]["beta"].append(betaM)
                    percents["fp1"]["theta"].append(thetaM)
                if fp == "fp2":
                    percents["fp2"]["alpha"].append(alphaM)
                    percents["fp2"]["beta"].append(betaM)
                    percents["fp2"]["theta"].append(thetaM)
                if fp == "fpz":
                    percents["fpz"]["alpha"].append(alphaM)
                    percents["fpz"]["beta"].append(betaM)
                    percents["fpz"]["theta"].append(thetaM)

        for i in patient_chanels:
            for j in patient_chanels[i]:
                patient_chanels[i][j] = abs(
                    int(np.mean(patient_chanels[i][j])))

        for i in percents:
            for j in percents[i]:
                percents[i][j] = abs(
                    np.mean(percents[i][j]))
        means[stage] = {"patients": patient_chanels,
                        "percents": percents,
                        "means": [abs(int(Mean(alpha_means))),
                                  abs(int(Mean(beta_means))),
                                  abs(int(Mean(theta_means)))]}
    return means
